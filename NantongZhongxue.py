#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Mar 21 15:06:17 2017

@author: wf
"""

#Nantong Zhongxue
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

# to wirte in excel file, seem complex
''' 
def write_to_excel(content):
    #新建一个workbook  
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    ws['A'+line_num] = content[1]
    ws['B'+line_num] = content[2]
    ws['C'+line_num] = content[3]
    ws['D'+line_num] = content[4]
    ws['E'+line_num] = content[5]
    ws['F'+line_num] = content[7]
    ws['G'+line_num] = content[9]
    wb.save("sample.xlsx")


write_to_excel()
'''
url = 'http://gklq.ntzx.cn/index.asp?page=1&jlh=0'


def Write_to_text(content):
    content.append('\n')
    with open ('Nanontzhongxue.txt','a') as f:
        f.write(' '.join(content))
        
        
def Fetch_and_write(url):
    response = requests.get(url)
    response.encoding = 'gb2312'
    soup = BeautifulSoup(response.text,'lxml')

    for i in soup.find_all('tr', bgcolor = ['#eeeeee', 'silver']):
        lst=[]
        for  child in i.descendants:
            if isinstance(child,str) and child != '\n':
                lst.append(child)
        lst.append('\n')
        with open ('Nanontzhongxue.txt','a') as f:
            f.write(' '.join(lst))
    # print(lst)
Fetch_and_write(url)  

